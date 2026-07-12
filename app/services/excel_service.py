from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook

# Excel 파일 입출력과 Excel 데이터 정리를 담당하는 서비스.
# 업로드된 Excel을 Python dict 리스트로 변환하고,
# 대사 결과를 다운로드 가능한 Excel 파일로 생성.
class ExcelService:
    def validate_excel_file(self, file: UploadFile) -> None:
        if not file.filename:
            raise ValueError("파일명이 없습니다.")

        if not file.filename.lower().endswith(".xlsx"):
            raise ValueError("xlsx 파일만 업로드할 수 있습니다.")

    async def read_excel(self, file: UploadFile, required_columns: list[str]) -> list[dict[str, Any]]:
        self.validate_excel_file(file)

        content = await file.read()
        #  업로드 파일을 디스크에 저장하지 않고 메모리에서 열게함.
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValueError("Excel 파일이 비어 있습니다.")

        headers = [self._normalize_header(value) for value in rows[0]]

        missing_columns = [
            column for column in required_columns
            if column not in headers
        ]

        if missing_columns:
            raise ValueError(f"필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}")

        result: list[dict[str, Any]] = []

        # 두번째 행부터 데이터로 처리. 첫번째 행은 헤더.
        for row_number, row in enumerate(rows[1:], start=2):
            if self._is_empty_row(row):
                continue

            row_data = dict(zip(headers, row, strict=False))
            row_data["row_number"] = row_number
            result.append(row_data)

        return result

    def create_result_excel(self, results: list[dict[str, Any]]) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "reconciliation_results"

        headers = [
            "item_identifier",
            "item_type",
            "institution_name",
            "currency",
            "base_date",
            "confirmation_amount",
            "book_amount",
            "difference_amount",
            "is_matched",
            "result_type",
            "message",
        ]

        sheet.append(headers)

        for result in results:
            sheet.append([
                result.get("item_identifier"),
                result.get("item_type"),
                result.get("institution_name"),
                result.get("currency"),
                result.get("base_date"),
                result.get("confirmation_amount"),
                result.get("book_amount"),
                result.get("difference_amount"),
                result.get("is_matched"),
                result.get("result_type"),
                result.get("message"),
            ])

        # 메모리 파일로 저장
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    def to_decimal(self, value: Any) -> Decimal | None:
        if value is None:
            return None

        text = str(value).strip()

        if text == "":
            return None

        text = text.replace(",", "")

        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    # identifier를 정규화하여 대사하기 쉽게 함 ex) 123-456 > 123456
    def normalize_identifier(self, value: Any) -> str:
        if value is None:
            return ""

        text = str(value).strip()
        text = text.replace("-", "")
        text = text.replace(" ", "")

        return text

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _is_empty_row(row: tuple[Any, ...]) -> bool:
        return all(
            value is None or str(value).strip() == ""
            for value in row
        )